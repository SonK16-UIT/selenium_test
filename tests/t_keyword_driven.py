# Updated Selenium Keyword-Driven Framework with Full Keyword Implementation

import pytest
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from tests.utils.selenium_utils import wait_for_element

class ReadExcel:
    def __init__(self, file_path, sheet_name):
        self.workbook = load_workbook(filename=file_path)
        self.sheet = self.workbook[sheet_name]

    def get_test_steps(self):
        return [row for row in self.sheet.iter_rows(min_row=2, values_only=True)]

class PageActions:
    def __init__(self, driver):
        self.driver = driver

    def navigate(self, url):
        self.driver.get(url)
        print(f"Navigated to {url}")

    def enterEmail(self, email):
        wait_for_element(self.driver, By.XPATH, "//input[@id='email']").send_keys(email)

    def enterPassword(self, password):
        wait_for_element(self.driver, By.XPATH, "//input[@id='password']").send_keys(password)

    def clickSignIn(self):
        wait_for_element(self.driver, By.XPATH, "//button[@id='login_button']").click()

    def verifyUserName(self, expected_name):
        actual = wait_for_element(self.driver, By.XPATH, "//p[contains(text(), 'SONK16789')]").text
        assert actual == expected_name, f'Expected {expected_name}, got {actual}'

    def clickDropDown(self):
        wait_for_element(self.driver, By.XPATH, "//svg[contains(@class,'svg[@class='MuiSvgIcon-root MuiSvgIcon-fontSizeMedium css-pscqdo')]").click()

    def logout(self):
        wait_for_element(self.driver, By.XPATH, "//li[contains(@class,'MuiButtonBase-root MuiMenuItem-root MuiMenuItem-gutters MuiMenuItem-root MuiMenuItem-gutters css-ovm7gc')]").click()

    def verifySignBtn(self, expected_text):
        actual = wait_for_element(self.driver, By.XPATH, "//a[@id='login_button']").text
        assert actual == expected_text, f'Expected {expected_text}, got {actual}'

@pytest.fixture(scope="module")
def driver_instance(request):
    from selenium import webdriver
    driver = webdriver.Chrome()
    driver.maximize_window()
    yield driver
    driver.quit()

@pytest.mark.parametrize("step_num, keyword, locator, value", ReadExcel('resource/DemoFile.xlsx', 'Keyword').get_test_steps())
def test_keyword_execution(driver_instance, step_num, keyword, locator, value):
    page = PageActions(driver_instance)
    try:
        method = getattr(page, keyword, None)
        assert method, f'Method {keyword} not found.'
        if value:
            method(value)
        else:
            method()
        print(f"Step {step_num}: {keyword} passed.")
    except Exception as e:
        pytest.fail(f"Step {step_num}: {keyword} failed with error: {e}")
